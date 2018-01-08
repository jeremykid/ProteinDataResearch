from openpyxl import load_workbook
import numpy as np
# import matplotlib.pyplot as plt
import pandas as pd
from sklearn import datasets, linear_model, metrics
from sklearn import preprocessing
from sklearn.model_selection import LeaveOneOut
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.preprocessing import normalize
from sklearn.neural_network import MLPClassifier, MLPRegressor
from sklearn import datasets, linear_model

# def read_data(fileName):
#   wb = load_workbook(filename = fileName)
#   sheet_ranges = wb["Sheet1"]["A4"].value

# def data_convert():
#   protineList = wb["Sheet1"]

# Todo get the  

def checkDistance(protineArray):
    minValue = np.amin(protineArray)
    maxValue = np.amax(protineArray)
    distance = 0.1*minValue

    if (maxValue - minValue >= distance):
        return 1
    else:
        return 0

def looFilter():
    loo = LeaveOneOut()
    # loo.get_n_splits(X)
    for train_index, test_index in loo.split(X_All):
        X_train, X_test = X_All[train_index], X_All[test_index]
        print(X_train, X_test)

def LogisticRegression(X_All, y_All):
    logistic = linear_model.LogisticRegression()
    kbest = SelectKBest(f_regression, k=15)
    X_All = kbest.fit_transform(X_All, y_All)

    loo = LeaveOneOut()
    loo.get_n_splits(X_All)
    for train_index, test_index in loo.split(X_All):
        # print test_index
        X_train, X_test = X_All[train_index], X_All[test_index]
        y_train, y_test = y_All[train_index], y_All[test_index]
        # print(X_train, X_test)
        y_ = logistic.fit(X_train, y_train).predict(X_test)
        print (y_)

    retained2 = kbest.get_support(True)
    logistic.fit(X_All, y_All)
    # print "logistic.densify().coef_"  

    # print logistic.densify().coef_.shape
    return logistic.densify().coef_, retained2

def multiLayerClassification(X_All, y_All):
    clf = MLPClassifier(solver='lbfgs', alpha=1e-5,
                     hidden_layer_sizes=(5, 2), random_state=1)
    kbest = SelectKBest(f_regression, k=15)
    X_All = kbest.fit_transform(X_All, y_All)

    loo = LeaveOneOut()
    loo.get_n_splits(X_All)
    for train_index, test_index in loo.split(X_All):
        # print test_index
        X_train, X_test = X_All[train_index], X_All[test_index]
        y_train, y_test = y_All[train_index], y_All[test_index]
        # print(X_train, X_test)
        y_ = clf.fit(X_train, y_train).predict(X_test)
        print (y_)


def multiLayerRegression(X_All, y_All):
    # clf = MLPRegressor(hidden_layer_sizes=(5, ), activation='relu', solver='adam', alpha=0.01, 
    #     batch_size='auto', learning_rate='constant', learning_rate_init=0.01, power_t=0.5, 
    #     max_iter=1, shuffle=True, random_state=1, tol=0.01, verbose=False, warm_start=False, 
    #     momentum=0.9, nesterovs_momentum=True, early_stopping=False, validation_fraction=0.1, beta_1=0.9, 
    #     beta_2=0.99, epsilon=1e-02)
    clf = MLPRegressor(hidden_layer_sizes=(5, ), activation='relu', verbose=True, learning_rate_init=1, learning_rate='adaptive', max_iter=500,)
    kbest = SelectKBest(f_regression, k=15)
    X_All = kbest.fit_transform(X_All, y_All)

    loo = LeaveOneOut()
    loo.get_n_splits(X_All)
    for train_index, test_index in loo.split(X_All):
        # print test_index
        X_train, X_test = X_All[train_index], X_All[test_index]
        y_train, y_test = y_All[train_index], y_All[test_index]
        # print(X_train, X_test)
        y_ = clf.partial_fit(X_train, y_train).predict(X_test)
        print (y_)
    

def linearRegression(X_All, y_All):
    # Create linear regression object
    regr = linear_model.LinearRegression()

    kbest = SelectKBest(f_regression, k=15)
    X_All = kbest.fit_transform(X_All, y_All)

    loo = LeaveOneOut()
    loo.get_n_splits(X_All)
    for train_index, test_index in loo.split(X_All):
        # print test_index
        X_train, X_test = X_All[train_index], X_All[test_index]
        y_train, y_test = y_All[train_index], y_All[test_index]
        # print(X_train, X_test)
        y_ = regr.fit(X_train, y_train).predict(X_test)
        print (y_)


def readY():
    wb = load_workbook(filename = 'saliva_spread_sheet.xlsx')
    sheet_ranges = wb["Sheet1"]
    # 28 57 22 23 33 43 52 72
    samplePostion = ["AH","BK","AB","AC","AL","AW","BF","BZ"]
    y = []
    for samplePostionIndex in range(len(samplePostion)):
        if (sheet_ranges[samplePostion[samplePostionIndex]+"99"].value == 1):
            y.append(1)
        else:
            y.append(0)   
    print y
    ## age ##
    # for samplePostionIndex in range(len(samplePostion)):
    #     #y.append(sheet_ranges[samplePostion[samplePostionIndex]+"6"].value)
    #     if (sheet_ranges[samplePostion[samplePostionIndex]+"6"].value >= 70):
    #         y.append(1)
    #     else:
    #         y.append(0)       
    # print ("age" , y)
    y = np.array(y)     
    return y

def readX():
    wb = load_workbook(filename = 'Proteomic_data.xlsm')
    sheet_ranges = wb["Sheet1"]

    startIndex = 3
    endIndex = 308

    protineList = []
    samplePostion = ["H","I","J","K","L","M","N","O"]


    for samplePostionIndex in range(len(samplePostion)):
        protineRow = []
        for i in range(startIndex,endIndex+1):

            protineRow.append(sheet_ranges[samplePostion[samplePostionIndex]+str(i)].value)
        protineList.append(protineRow)

    ## READ ## 

    protineList = np.array(protineList)
    return protineList

def readProtineName():
    wb = load_workbook(filename = 'Proteomic_data.xlsm')
    sheet_ranges = wb["Sheet1"]

    startIndex = 3
    endIndex = 308

    protineNameList = {}


    for i in range(startIndex,endIndex+1):

        protineNameList[i-3] = sheet_ranges['A'+str(i)].value

    ## READ ## 
    return protineNameList

def processCoefToProtineType(CoefArray, protineNameList, selectedFeatures):
    #sortedCoefArray = np.sort\(CoefArray\)
    positiveCoefArray = []
    negativeCoefArray = []

    for i in CoefArray[0]:
        if i >= 0:
            positiveCoefArray.append(i)
        else:
            negativeCoefArray.append(i)

    sortedPositiveCoefArray = np.sort(positiveCoefArray)[::-1]
    sortedNegativeCoefArray = np.sort(negativeCoefArray)

    print "---------- positive -------"
    for i in sortedPositiveCoefArray:
         
        #get Index by value
        #get protein name by index
        index = np.argwhere(CoefArray == i)

        name = protineNameList[selectedFeatures[index.item(1)]]
        #(CoefArray.where(i))
        print (name, i) #the index

    print "---------- negative --------"
    for i in sortedNegativeCoefArray:
    
        #get Index by value
        #get protein name by index
        index = np.argwhere(CoefArray == i)
        name = protineNameList[selectedFeatures[index.item(1)]]
        #(CoefArray.where(i))
        print (name, i) #the index

def normalization(X):
    return normalize(X, axis=0, norm='l2')

def getDrugData():
    wb = load_workbook(filename = 'drug_data.xlsx')
    sheet_ranges = wb["Sheet1"]
    # 28 57 22 23 33 43 52 72
    samplePostion = ["D","H","B","C","E","F","G","I"]
    y = []
    for samplePostionIndex in range(len(samplePostion)):
        # 4 - 32
        #print (sheet_ranges[samplePostion[samplePostionIndex]+"4"].value)

        y.append(float(sheet_ranges[samplePostion[samplePostionIndex]+"33"].value))

        #class
        # if (sheet_ranges[samplePostion[samplePostionIndex]+"3"].value == 1):
        #     y.append(1)
        # else:
        #     y.append(0)   
        #end of class
    print y
    ## age ##
    # for samplePostionIndex in range(len(samplePostion)):
    #     #y.append(sheet_ranges[samplePostion[samplePostionIndex]+"6"].value)
    #     if (sheet_ranges[samplePostion[samplePostionIndex]+"6"].value >= 70):
    #         y.append(1)
    #     else:
    #         y.append(0)       
    # print ("age" , y)
    y = np.array(y)     
    return y


def main():

    protineNameList = readProtineName()
    #### READ FROM ORIGINAL FILE ####
    y = readY()

    #### READ FROM DRUG DATA ####
    #y = getDrugData()
    X = readX()
    
    X = normalization(X)

    #y = preprocessing.MinMaxScaler().fit_transform(y)
    #print (y)
    #multiLayerRegression(X,y)
    multiLayerClassification(X,y)
    #linearRegression(X,y)
    return
    coefList, selectedFeatures = LogisticRegression(X,y)
    
    processCoefToProtineType(coefList, protineNameList, selectedFeatures)
    #protineListT = protineList.T

    # value0 = 0
    # value1 = 0
    # # print (protineListT[208,:])
    # for i in protineListT:
    #     if checkDistance(i):
    #         value1 += 1
    #     else:
    #         value0 += 1
    # print (value1,value0)
main()


